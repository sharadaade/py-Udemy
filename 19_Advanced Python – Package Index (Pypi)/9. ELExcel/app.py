# *****************---------------- Working with Excel Spreadsheets

import openpyxl

# work_book = openpyxl.Workbook()

work_book = openpyxl.load_workbook("users.xlsx")
# print(work_book.sheetnames)

sheet = work_book["Sheet_1"]
# print(sheet)

# --------------------------------------------------------------
# *-*-*-*-*-* Accessing individual cells -----------------------

# Approach one ->>>> *-*- cell coordiantes --------
# cell = sheet["A1"]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)

# Approach two ->>>> *-*- cell() method --------
# cell = sheet.cell(row=1, column=1)
# print(cell)

# ----------------------------------------------------------------
# *-*-*-*-*-* Accessing rows/columns cells -----------------------

# print("Number of Rows:", sheet.max_row)
# print("Number of Columns:", sheet.max_column)

# ---------------------------------------------------------------------
# *-*-*-*-*-* Iterating over rows/columns cells -----------------------

# for row in range(1, sheet.max_row + 1):
#     for col in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, col)
#         print(cell.value)

# for col in range(1, sheet.max_column + 1):
#     for row in range(1, sheet.max_row + 1):
#         cell = sheet.cell(row, col)
#         print(cell.value)

# --------------------------------------------------------------
# *-*-*-*-*-* Accessing a range of cells -----------------------


# *-*- getting all the cells in a col --------
# column = sheet["A"]
# print(column)

# *-*- getting a range of cells --------
# print(sheet["A1:B5"])

# *-*- getting a range of cells using coodinates --------
# print(sheet["A:C"])

# *-*- getting all the cells in the first row --------
# print(sheet[1])

# *-*- getting all the cells in a range of rows --------
# print(sheet[1:3])

# --------------------------------------------------------------
# *-*-*-*-*-* sheet object methods -----------------------------

# *-*- append()
# sheet.append(["wolf", 555, "online"])

# *-*- insert_rows(starting index, # of empty rows)
# sheet.insert_rows(3, 10)

# *-*- insert_cols(starting index, # of empty cols)
# sheet.insert_cols(2, 5)

# *-*- delete_rows(starting index, # of rows)
# sheet.delete_rows(3, 10)

# *-*- delete_cols(starting index, # of cols)
# sheet.delete_cols(2, 5)


# work_book.save("users.xlsx")


# --------------------------------------------------------------
# *-*-*-*-*-* adding sheets to the workbook -----------------------------
work_book.create_sheet("PY_Sheet_1")
work_book.save("new_users.xlsx")
